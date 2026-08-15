#!/usr/bin/env python3
"""Build the public Nebraska budget dashboard data file from official sources.

The scraper discovers documents from each agency's index page instead of guessing
filenames. Every PDF/XLSX download is signature-checked before parsing, and a
failed source keeps the last known-good section of the JSON file.
"""

from __future__ import annotations

import argparse
import calendar
import json
import re
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urljoin
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DAS_REPORTS_URL = "https://das.nebraska.gov/accounting/financial_reports.php"
LEGISLATURE_REPORTS_URL = "https://nebraskalegislature.gov/reports/fiscal.php"
CURRENT_GF_STATUS_URL = "https://nebraskalegislature.gov/FloorDocs/Current/PDF/Budget/status.pdf"
REVENUE_REPORTS_URL = (
    "https://revenue.nebraska.gov/about/news-releases/"
    "general-fund-receipts-news-releases"
)
REVENUE_FORECASTS_PATH = Path(__file__).with_name("revenue_forecasts.json")
AGENCY_BUDGET_URL = "https://statespending.nebraska.gov/CurrentFiscalYearBudget"
CENSUS_POPULATION_URL = "https://www.census.gov/quickfacts/fact/table/NE/PST045225"
NEBRASKA_POPULATION = 2_018_006
NEBRASKA_POPULATION_AS_OF = "July 1, 2025 estimate"
USER_AGENT = "NebraskaBudgetDashboard/2.0 (+https://github.com/tpendrell/nebraska-budget-dashboard-v2)"


class SourceError(RuntimeError):
    """An official source could not be safely fetched or parsed."""


@dataclass(frozen=True)
class Link:
    url: str
    text: str


class LinkParser(HTMLParser):
    def __init__(self, base_url: str):
        super().__init__()
        self.base_url = base_url
        self.links: list[Link] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._text = []

    def handle_data(self, data):
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "a" and self._href is not None:
            text = re.sub(r"\s+", " ", " ".join(self._text)).strip()
            self.links.append(Link(urljoin(self.base_url, self._href), text))
            self._href = None
            self._text = []


class TableParser(HTMLParser):
    """Small HTML table extractor; keeps all table cells as plain text."""

    def __init__(self):
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table":
            self._table = []
        elif tag == "tr" and self._table is not None:
            self._row = []
        elif tag in ("th", "td") and self._row is not None:
            self._cell = []

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("th", "td") and self._cell is not None and self._row is not None:
            self._row.append(re.sub(r"\s+", " ", " ".join(self._cell)).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None and self._table is not None:
            if any(self._row):
                self._table.append(self._row)
            self._row = None
        elif tag == "table" and self._table is not None:
            if self._table:
                self.tables.append(self._table)
            self._table = None


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def fetch_bytes(url: str, timeout: int = 60) -> tuple[bytes, str]:
    request = Request(quote(url, safe=":/?&=%+#"), headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    try:
        with urlopen(request, timeout=timeout) as response:
            body = response.read()
            content_type = response.headers.get_content_type()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise SourceError(f"Could not fetch {url}: {exc}") from exc
    if not body:
        raise SourceError(f"Official source returned an empty response: {url}")
    return body, content_type


def fetch_html(url: str) -> str:
    body, content_type = fetch_bytes(url)
    if body.lstrip().startswith((b"%PDF", b"PK\x03\x04")):
        raise SourceError(f"Expected an HTML index page but received a document: {url}")
    if "html" not in content_type and b"<html" not in body[:1000].lower():
        raise SourceError(f"Expected HTML from {url}; received {content_type}")
    return body.decode("utf-8", errors="replace")


def official_links(url: str) -> tuple[str, list[Link]]:
    html = fetch_html(url)
    parser = LinkParser(url)
    parser.feed(html)
    return html, parser.links


def download_document(url: str, destination: Path, expected: str) -> Path:
    body, content_type = fetch_bytes(url)
    expected = expected.lower()
    if expected == "pdf" and not body.lstrip().startswith(b"%PDF"):
        raise SourceError(f"{url} returned {content_type}, not a PDF. Refusing to parse it.")
    if expected == "xlsx" and not body.startswith(b"PK\x03\x04"):
        raise SourceError(f"{url} returned {content_type}, not an XLSX workbook. Refusing to parse it.")
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(body)
    return destination


def parse_target_month(value: str | None) -> date | None:
    if not value:
        return None
    try:
        parsed = datetime.strptime(value, "%Y-%m").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError("--month must use YYYY-MM") from exc
    return parsed.replace(day=calendar.monthrange(parsed.year, parsed.month)[1])


def fiscal_period_to_calendar(fiscal_year: int, period: int) -> date:
    if not 1 <= period <= 12:
        raise ValueError(f"invalid fiscal period: {period}")
    year = fiscal_year - 1 if period <= 6 else fiscal_year
    month = period + 6 if period <= 6 else period - 6
    return date(year, month, calendar.monthrange(year, month)[1])


def fiscal_year_label(day: date | None = None) -> str:
    day = day or date.today()
    start = day.year if day.month >= 7 else day.year - 1
    return f"FY{start}-{str(start + 1)[-2:]}"


def _number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("() ").replace("−", "-")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -abs(number) if negative else number


def _amounts(line: str) -> list[int]:
    values = []
    for match in re.finditer(r"(?<![\w.])\(?-?\$?\d[\d,]*(?:\.\d+)?\)?(?![%\w])", line):
        token = match.group(0)
        digits = re.sub(r"\D", "", token)
        if "," not in token and "$" not in token and len(digits) < 5:
            continue
        values.append(round(_number(token)))
    return values


def _row_amounts(line: str) -> list[int]:
    """Extract table cells, preserving published dash placeholders as zeroes."""
    values = []
    pattern = r"(?<![\w.])(?:--+|—+|\(?-?\$?\d[\d,]*(?:\.\d+)?\)?)(?![%\w.])"
    for match in re.finditer(pattern, line):
        token = match.group(0)
        values.append(0 if re.fullmatch(r"--+|—+", token) else round(_number(token)))
    return values


def _pdf_to_text(path: Path | None) -> str:
    if path is None:
        raise SourceError("No PDF source was discovered")
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            check=True,
            capture_output=True,
            text=True,
            timeout=120,
        )
    except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        raise SourceError(f"Could not extract text from {path.name}: {exc}") from exc
    return result.stdout


def _source(name: str, url: str, period: str, **extra) -> dict:
    return {"name": name, "url": url, "period": period, **extra}


# DAS Operating Investment Pool

OIP_RE = re.compile(r"Operating[_-]Investment[_-]Pool.*?_(\d{4})-(\d{2})\.(xlsx|pdf)$", re.I)


def discover_oip(links: Iterable[Link], target: date | None = None) -> tuple[date, dict[str, str]]:
    reports: dict[date, dict[str, str]] = {}
    for link in links:
        match = OIP_RE.search(link.url)
        if not match:
            continue
        report_date = fiscal_period_to_calendar(int(match.group(1)), int(match.group(2)))
        reports.setdefault(report_date, {})[match.group(3).lower()] = link.url
    candidates = [day for day in reports if target is None or day <= target]
    if not candidates:
        raise SourceError("No OIP Excel/PDF links were found on the DAS reports page")
    newest = max(candidates)
    return newest, reports[newest]


def fetch_oip(work_dir: Path, target: date | None = None) -> tuple[Path, date, str]:
    _, links = official_links(DAS_REPORTS_URL)
    report_date, documents = discover_oip(links, target)
    errors = []
    for extension in ("xlsx", "pdf"):
        if extension not in documents:
            continue
        try:
            path = download_document(
                documents[extension], work_dir / f"oip-{report_date.isoformat()}.{extension}", extension
            )
            return path, report_date, documents[extension]
        except SourceError as exc:
            errors.append(str(exc))
    raise SourceError("; ".join(errors) or "The latest OIP report had no downloadable document")


def _normalized_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def parse_oip_xlsx(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    header_row = None
    headers: dict[int, str] = {}
    interest_rate = None

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        normalized = [_normalized_header(value) for value in row]
        joined = " | ".join(normalized)
        raw_joined = " | ".join(str(value or "") for value in row)
        rate_match = re.search(r"interest rate[^0-9]*(\d+(?:\.\d+)?)", raw_joined, re.I)
        if rate_match:
            interest_rate = float(rate_match.group(1))
            if interest_rate <= 1:
                interest_rate *= 100
        if "fund" in joined and "average daily balance" in joined and "allocated interest" in joined:
            header_row = row_index
            headers = {index: value for index, value in enumerate(normalized)}
            break

    if header_row is None:
        raise SourceError("OIP workbook does not contain the expected named columns")

    def find_column(*needles: str) -> int:
        for index, header in headers.items():
            if all(needle in header for needle in needles):
                return index
        raise SourceError(f"OIP workbook is missing column: {' '.join(needles)}")

    fund_col = find_column("fund")
    title_col = next(
        (i for i, h in headers.items() if "fund" in h and any(x in h for x in ("name", "title", "description"))),
        None,
    )
    balance_col = find_column("average", "daily", "balance")
    interest_col = find_column("allocated", "interest")

    funds = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_id = row[fund_col] if fund_col < len(row) else None
        if raw_id is None:
            continue
        if isinstance(raw_id, (int, float)):
            fund_id = str(int(raw_id)).zfill(5)
        else:
            match = re.fullmatch(r"\s*(\d{4,5})(?:\.0)?\s*", str(raw_id))
            if not match:
                continue
            fund_id = match.group(1).zfill(5)
        title = str(row[title_col] or "").strip() if title_col is not None and title_col < len(row) else ""
        balance = _number(row[balance_col] if balance_col < len(row) else 0)
        interest = _number(row[interest_col] if interest_col < len(row) else 0)
        funds.append({
            "id": fund_id,
            "title": title or f"Fund {fund_id}",
            "balance": round(balance, 2),
            "interest": round(interest, 2),
            "hasOipEntry": True,
        })

    if len(funds) < 10:
        raise SourceError(f"OIP workbook yielded only {len(funds)} fund rows")

    if interest_rate is None:
        for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
            text = " ".join(str(value or "") for value in row)
            match = re.search(r"Interest\s*Rate[^0-9]*(\d+(?:\.\d+)?)\s*%?", text, re.I)
            if match:
                interest_rate = float(match.group(1))
                if interest_rate <= 1:
                    interest_rate *= 100
                break

    return {
        "funds": sorted(funds, key=lambda item: item["balance"], reverse=True),
        "macro": {
            "totalBalance": round(sum(fund["balance"] for fund in funds), 2),
            "totalInterest": round(sum(fund["interest"] for fund in funds), 2),
            "interestRate": interest_rate,
            "effectiveYield": f"{interest_rate:.5f}%" if interest_rate is not None else "Not published",
            "reportedFunds": len(funds),
            "activeFunds": sum(1 for fund in funds if fund["balance"] != 0),
        },
    }


def parse_oip_pdf(path: Path) -> dict:
    text = _pdf_to_text(path)
    rate_match = re.search(r"Interest\s*Rate[^0-9]*(\d+(?:\.\d+)?)\s*%", text, re.I)
    interest_rate = float(rate_match.group(1)) if rate_match else None
    funds = []
    for line in text.splitlines():
        match = re.match(
            r"^\s*(\d{5})\s+(.+?)\s+(\(?-?\$?[\d,]+(?:\.\d+)?\)?)\s+"
            r"(\(?-?\$?[\d,]+(?:\.\d+)?\)?)(?:\s+\d+)?\s*$",
            line,
        )
        if not match:
            continue
        funds.append({
            "id": match.group(1),
            "title": re.sub(r"\s+", " ", match.group(2)).strip(),
            "balance": round(_number(match.group(3)), 2),
            "interest": round(_number(match.group(4)), 2),
            "hasOipEntry": True,
        })
    if len(funds) < 10:
        raise SourceError(f"OIP PDF yielded only {len(funds)} fund rows")
    return {
        "funds": sorted(funds, key=lambda item: item["balance"], reverse=True),
        "macro": {
            "totalBalance": round(sum(x["balance"] for x in funds), 2),
            "totalInterest": round(sum(x["interest"] for x in funds), 2),
            "interestRate": interest_rate,
            "effectiveYield": f"{interest_rate:.5f}%" if interest_rate is not None else "Not published",
            "reportedFunds": len(funds),
            "activeFunds": sum(1 for x in funds if x["balance"] != 0),
        },
    }


def parse_oip(path: Path) -> dict:
    return parse_oip_xlsx(path) if path.suffix.lower() == ".xlsx" else parse_oip_pdf(path)


def merge_fund_history(current_funds: list[dict], previous_funds: list[dict], report_date: date) -> list[dict]:
    """Attach a compact monthly OIP history without inventing missing periods.

    The OIP publishes average daily balances, not point-in-time cash balances. A
    repeated weekly refresh for the same report month replaces that month's
    point; a newly published month appends one point and calculates the change
    from the prior published month. Twenty-four months are retained.
    """
    previous_by_id = {str(fund.get("id")): fund for fund in previous_funds or []}
    period = report_date.strftime("%Y-%m")
    label = report_date.strftime("%b %Y")
    merged: list[dict] = []

    for fund in current_funds:
        item = dict(fund)
        prior = previous_by_id.get(str(item.get("id")), {})
        history = []
        for point in prior.get("history", []) or []:
            point_period = str(point.get("period") or "")
            point_balance = point.get("balance")
            if point_balance is None and point.get("b") is not None:
                point_balance = float(point["b"]) * 1_000_000
            if not point_period or point_balance is None:
                continue
            history.append({
                "period": point_period,
                "label": point.get("label") or point.get("m") or point_period,
                "balance": round(float(point_balance), 2),
                "interest": round(float(point.get("interest") or 0), 2),
            })

        current_point = {
            "period": period,
            "label": label,
            "balance": round(float(item.get("balance") or 0), 2),
            "interest": round(float(item.get("interest") or 0), 2),
        }
        history = [point for point in history if point["period"] != period]
        history.append(current_point)
        history = sorted(history, key=lambda point: point["period"])[-24:]

        item["history"] = history
        item["delta"] = round(history[-1]["balance"] - history[-2]["balance"], 2) if len(history) > 1 else 0
        merged.append(item)

    return merged


# Legislature reports

def discover_legislature_documents(links: Iterable[Link]) -> dict:
    pdfs = [link for link in links if ".pdf" in link.url.lower()]
    status = next((link for link in pdfs if "current general fund status" in link.text.lower()), None)
    budget_candidates = []
    for link in pdfs:
        combined = f"{link.text} {link.url}"
        if "budget" not in combined.lower() and "final" not in combined.lower():
            continue
        years = [int(value) for value in re.findall(r"20\d{2}", combined)]
        if years:
            budget_candidates.append((max(years), link))
    budget = max(budget_candidates, default=(0, None), key=lambda item: item[0])[1]
    directory_candidates = []
    for link in pdfs:
        combined = f"{link.text} {link.url}"
        if "directory" not in combined.lower():
            continue
        years = [int(value) for value in re.findall(r"20\d{2}", combined)]
        directory_candidates.append((max(years, default=0), link))
    latest_directory_year = max((year for year, _ in directory_candidates), default=0)
    directories = [link for year, link in directory_candidates if year == latest_directory_year]
    if not status and not budget:
        raise SourceError("The Legislature fiscal reports page did not expose a GF status or budget PDF")
    return {"status": status, "budget": budget, "directories": directories}


def _find_table_row(text: str, label_pattern: str, minimum_values: int = 4) -> list[int]:
    for line in text.splitlines():
        label = re.search(label_pattern, line, re.I)
        if not label:
            continue
        values = _row_amounts(line[label.end():])
        if len(values) >= minimum_values:
            return values
    return []


def parse_gf_status_text(text: str, as_of: date | None = None) -> dict:
    anchors = [match.start() for match in re.finditer(r"General Fund Financial Status", text, re.I)]
    candidates = [text[start:start + 12000] for start in anchors] or [text]
    selected = None
    years: list[str] = []
    row_specs = [
        ("Unobligated Beginning Balance", r"Unobligated\s+Beginning\s+Balance"),
        ("General Fund Net Revenues", r"General\s+Fund\s+Net\s+Revenues"),
        ("General Fund Appropriations", r"General\s+Fund\s+Appropriations"),
        ("Ending Balance", r"Ending\s+Balance"),
        ("Minimum Reserve at 3%", r"Min(?:imum)?\.?\s+Reserve(?:\s+at)?\s+3(?:\.0)?%"),
        ("Excess / (Shortfall)", r"Excess\s*/?\s*\(?Shortfall\)?"),
    ]
    for section in candidates:
        year_matches = re.findall(r"FY\s*((?:20)?\d{2})\s*[-–/]\s*(\d{2,4})", section, re.I)
        found_years = []
        for start, end in year_matches:
            start_year = int(start) if len(start) == 4 else 2000 + int(start)
            label = f"FY{start_year}-{end[-2:]}"
            if label not in found_years:
                found_years.append(label)
        if len(found_years) < 4:
            continue
        rows = []
        for label, pattern in row_specs:
            sparse_biennium_row = label in {"Minimum Reserve at 3%", "Excess / (Shortfall)"}
            values = _find_table_row(section, pattern, 2 if sparse_biennium_row else len(found_years))
            if values:
                used_years = found_years if len(values) >= len(found_years) else found_years[-len(values):]
                rows.append({"label": label, "values": dict(zip(used_years, values[:len(used_years)]))})
        if len(rows) >= 4:
            # The official one-page status centers biennium-only reserve cells
            # between two fiscal-year columns. pdftotext places those values in
            # the first column of each pair; move them to the biennium-ending
            # year used by the dashboard.
            for item in rows:
                if item["label"] not in {"Minimum Reserve at 3%", "Excess / (Shortfall)"}:
                    continue
                for index in range(2, len(found_years), 2):
                    prior_year = found_years[index - 1]
                    ending_year = found_years[index]
                    if item["values"].get(ending_year, 0) == 0 and item["values"].get(prior_year, 0) != 0:
                        item["values"][ending_year] = item["values"][prior_year]
                        item["values"][prior_year] = 0
            selected = rows
            years = found_years
            break
    if not selected:
        raise SourceError("Could not locate a complete General Fund Financial Status table")

    def row(label: str) -> dict:
        return next((item["values"] for item in selected if item["label"] == label), {})

    appropriations = row("General Fund Appropriations")
    revenues = row("General Fund Net Revenues")
    if not all(3_000_000_000 < abs(value) < 9_000_000_000 for value in list(appropriations.values())[:3]):
        raise SourceError("GF status appropriation values failed range checks")
    if not all(3_000_000_000 < abs(value) < 9_000_000_000 for value in list(revenues.values())[:3]):
        raise SourceError("GF status revenue values failed range checks")

    current_fy = fiscal_year_label(as_of)
    if current_fy not in years:
        current_fy = years[min(1, len(years) - 1)]
    current_index = years.index(current_fy)
    biennium_index = current_index if current_index % 2 == 0 else min(current_index + 1, len(years) - 1)
    following_index = min(biennium_index + 2, len(years) - 1)
    variance = row("Excess / (Shortfall)")
    reserve_match = re.search(
        r"Cash Reserve Fund.{0,250}?projected.{0,250}?ending balance[^$\d]*"
        r"(?:\$)?([\d,.]+)\s*(million|billion)",
        text,
        re.I | re.S,
    )
    projected_reserve = 0
    if reserve_match:
        projected_reserve = _number(reserve_match.group(1))
        if (reserve_match.group(2) or "").lower() == "million":
            projected_reserve *= 1_000_000
        elif (reserve_match.group(2) or "").lower() == "billion":
            projected_reserve *= 1_000_000_000
    if not projected_reserve:
        reserve_values = _find_table_row(
            text,
            r"Projected\s+Unobligated\s+Ending\s+Balance",
            len(years),
        )
        if reserve_values:
            projected_reserve = reserve_values[years.index(current_fy)]
    status = {
        "fiscalYear": current_fy,
        "beginningBalance": row("Unobligated Beginning Balance").get(current_fy, 0),
        "netRevenues": revenues.get(current_fy, 0),
        "appropriations": appropriations.get(current_fy, 0),
        "endingBalance": row("Ending Balance").get(current_fy, 0),
        "currentBienniumFiscalYear": years[biennium_index],
        "minimumReserveVariance": variance.get(years[biennium_index], 0),
        "followingBienniumFiscalYear": years[following_index],
        "followingBienniumVariance": variance.get(years[following_index], 0),
        "cashReserveProjectedEndingBalance": round(projected_reserve),
    }
    return {"status": status, "table": selected, "years": years}


def parse_gf_status(path: Path, as_of: date | None = None) -> dict:
    return parse_gf_status_text(_pdf_to_text(path), as_of)


def parse_nefab_forecasts(path: Path) -> list[dict]:
    text = _pdf_to_text(path)
    match = re.search(
        r"Table\s+9\s*[-–]?\s*Actual and Projected General Fund Revenues"
        r"(.*?)(?=Table\s+10|Adjusted General Fund Revenues)",
        text,
        re.I | re.S,
    )
    if not match:
        return []
    section = match.group(1)
    year_matches = re.findall(r"FY\s*((?:20)?\d{2})\s*[-–/]\s*(\d{2,4})", section, re.I)
    years = []
    for start, end in year_matches:
        start_year = int(start) if len(start) == 4 else 2000 + int(start)
        label = f"FY{start_year}-{end[-2:]}"
        if label not in years:
            years.append(label)
    categories = [
        ("Sales & Use", r"Sales\s+(?:and|&)\s+Use\s+Tax"),
        ("Individual Income", r"Individual\s+Income\s+Tax"),
        ("Corporate Income", r"Corporate\s+Income\s+Tax"),
        ("Miscellaneous", r"Miscellaneous(?:\s+receipts)?"),
    ]
    result = []
    for name, pattern in categories:
        values = _find_table_row(section, pattern, min(4, len(years)))
        if len(values) < 4:
            continue
        used_years = years[-len(values):] if len(years) >= len(values) else [f"Column {i + 1}" for i in range(len(values))]
        result.append({"name": name, "values": dict(zip(used_years, values))})
    return result


def parse_lfo_directory(paths: Iterable[Path]) -> dict:
    descriptions = {
        "10000": {"title": "General Fund", "description": "The primary operating fund of the State."},
        "11000": {"title": "Cash Reserve Fund", "description": "Nebraska's rainy day fund."},
    }
    for path in paths:
        try:
            text = _pdf_to_text(path)
        except SourceError:
            continue
        for page in text.split("\f"):
            fund_match = re.search(r"FUND\s*:?\s*(\d{5})[\s:\-]+([^\n]+)", page, re.I)
            if not fund_match:
                continue
            fund_id = fund_match.group(1)
            description = re.search(
                r"PERMITTED USES\s*:?\s*(.+?)(?=\n\s*FUND SUMMARY|\n\s*REVENUE|\Z)",
                page, re.I | re.S,
            )
            authority = re.search(
                r"STATUTORY AUTHORITY\s*:?\s*(.+?)(?=\n\s*REVENUE|\n\s*PERMITTED|\Z)",
                page, re.I | re.S,
            )
            agency = re.search(r"AGENCY\s*:?\s*(?:#?\d+)?[\s\-:]*([^\n]+)", page, re.I)
            program = re.search(r"PROGRAM\s*:?\s*(?:#?\d+)?[\s\-:]*([^\n]+)", page, re.I)
            descriptions[fund_id] = {
                "title": fund_match.group(2).strip(),
                "description": re.sub(r"\s+", " ", description.group(1)).strip() if description else "",
                "statutory_authority": re.sub(r"\s+", " ", authority.group(1)).strip() if authority else "",
                "agency_name": agency.group(1).strip() if agency else "",
                "program": program.group(1).strip() if program else "",
            }
    return descriptions


# Department of Revenue monthly receipts

MONTHS = {name.lower(): index for index, name in enumerate(calendar.month_name) if name}
FISCAL_MONTHS = [
    "July", "August", "September", "October", "November", "December",
    "January", "February", "March", "April", "May", "June",
]
REVENUE_RE = re.compile(r"General_Fund_Receipts_News_Release_([A-Za-z]+)_(20\d{2}).*\.pdf$", re.I)


def discover_revenue_release(links: Iterable[Link], target: date | None = None) -> tuple[date, str]:
    reports = []
    for link in links:
        match = REVENUE_RE.search(link.url)
        if not match or match.group(1).lower() not in MONTHS:
            continue
        month = MONTHS[match.group(1).lower()]
        report_date = date(int(match.group(2)), month, calendar.monthrange(int(match.group(2)), month)[1])
        if target is None or report_date <= target:
            reports.append((report_date, link.url))
    if not reports:
        raise SourceError("No monthly General Fund receipts PDF was found on the Revenue index page")
    return max(reports, key=lambda item: item[0])


def revenue_fiscal_year(period: str) -> str:
    match = re.fullmatch(r"\s*([A-Za-z]+)\s+(20\d{2})\s*", period)
    if not match or match.group(1).lower() not in MONTHS:
        return ""
    month = MONTHS[match.group(1).lower()]
    year = int(match.group(2))
    start_year = year if month >= 7 else year - 1
    return f"FY{start_year}-{str(start_year + 1)[-2:]}"


def load_revenue_forecast(fiscal_year: str, path: Path = REVENUE_FORECASTS_PATH) -> dict:
    if not fiscal_year or not path.exists():
        return {}
    try:
        forecasts = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SourceError(f"Monthly revenue forecast file could not be read: {exc}") from exc
    schedule = forecasts.get(fiscal_year, {})
    rows = schedule.get("monthlyNetReceipts", [])
    months = [row.get("month") for row in rows]
    if schedule and months != FISCAL_MONTHS:
        raise SourceError(f"Monthly revenue forecast for {fiscal_year} must contain July through June in order")
    if any(not isinstance(row.get("forecast"), (int, float)) or row["forecast"] < 0 for row in rows):
        raise SourceError(f"Monthly revenue forecast for {fiscal_year} contains an invalid amount")
    return schedule


def merge_monthly_revenue_series(current: dict, previous: dict | None = None) -> dict:
    """Preload the certified forecast and retain actuals as monthly releases arrive."""
    fiscal_year = revenue_fiscal_year(current.get("period", ""))
    current["fiscalYear"] = fiscal_year
    schedule = load_revenue_forecast(fiscal_year)
    rows = [
        {"month": row["month"], "actual": None, "forecast": round(row["forecast"])}
        for row in schedule.get("monthlyNetReceipts", [])
    ]
    if not rows:
        rows = [
            {
                "month": row.get("month", ""),
                "actual": row.get("actual"),
                "forecast": round(row.get("forecast", 0)),
            }
            for row in current.get("monthlySeries", [])
        ]
    by_month = {row["month"]: row for row in rows if row.get("month")}

    previous = previous or {}
    previous_fiscal_year = previous.get("fiscalYear") or revenue_fiscal_year(previous.get("period", ""))
    if previous_fiscal_year == fiscal_year:
        for row in previous.get("monthlySeries", []):
            month = row.get("month")
            if month in by_month and row.get("actual") is not None:
                by_month[month]["actual"] = round(row["actual"])

    for row in current.get("monthlySeries", []):
        month = row.get("month")
        if not month:
            continue
        if month not in by_month:
            by_month[month] = {"month": month, "actual": None, "forecast": 0}
            rows.append(by_month[month])
        if row.get("actual") is not None:
            by_month[month]["actual"] = round(row["actual"])
        if row.get("forecast") is not None:
            by_month[month]["forecast"] = round(row["forecast"])

    current["monthlySeries"] = rows
    if schedule:
        current["monthlyForecast"] = {
            "basis": schedule.get("basis", ""),
            "certified": schedule.get("certified", ""),
            "source": schedule.get("source", ""),
        }
    return current


def parse_revenue_text(text: str, period: str) -> dict:
    candidates = []
    for page in text.split("\f"):
        for match in re.finditer(r"^\s*Net\s+Receipts\s*:?\s*$", page, re.I | re.M):
            candidates.append((page, match.start()))
    if not candidates:
        raise SourceError("Revenue release did not contain a Net Receipts table")
    # Current releases place the forecast table before a second table comparing
    # the current year with the previous year. Select the page whose headings
    # identify projected/forecast values instead of blindly using the last table.
    selected_page, selected_start = next(
        ((page, start) for page, start in candidates if re.search(r"Projected|Forecast", page, re.I)),
        candidates[-1],
    )
    section = selected_page[selected_start:]

    def row(pattern: str) -> list[int]:
        return _find_table_row(section, pattern, 2)

    def actual_forecast(values: list[int]) -> tuple[int, int]:
        return (values[3], values[4]) if len(values) >= 5 else (values[0], values[1])

    total_values = row(r"Total\s+Net(?:\s+Receipts)?")
    if len(total_values) < 2:
        raise SourceError("Revenue release Net Receipts total row was not parseable")
    ytd_actual, ytd_forecast = actual_forecast(total_values)
    categories = []
    for name, pattern in [
        ("Sales & Use", r"Sales\s+(?:and|&)\s+Use\s+Tax"),
        ("Individual Income", r"(?:Individual|Ind)\s+Income\s+Tax"),
        ("Corporate Income", r"(?:Corporate|Corp)\s+Income\s+Tax"),
        ("Miscellaneous", r"Misc(?:ellaneous)?(?:\s+Taxes)?"),
    ]:
        values = row(pattern)
        if len(values) >= 2:
            actual, forecast = actual_forecast(values)
            categories.append({"name": name, "actual": actual, "forecast": forecast})
    basis_match = re.search(
        r"((?:January|February|March|April|May|June|July|August|September|October|November|December)"
        r"\s+\d{1,2},\s+20\d{2})\s+(?:NEFAB\s+)?forecast",
        text,
        re.I,
    )
    if not basis_match:
        basis_match = re.search(
            r"forecast.{0,160}?\bon\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)"
            r"\s+\d{1,2},\s+20\d{2})",
            text,
            re.I | re.S,
        )
    return {
        "period": period,
        "ytdActual": ytd_actual,
        "ytdForecast": ytd_forecast,
        "categories": categories,
        "monthlySeries": [{"month": period.split()[0], "actual": total_values[0], "forecast": total_values[1]}],
        "nefabBasis": basis_match.group(1) if basis_match else "",
        "nefabForecasts": [],
    }


# Current fiscal-year agency appropriations

def _parse_agency_budget_spans(source: str) -> list[dict]:
    """Parse the State Spending site's current div/span-based agency listing."""
    rows: dict[int, dict[str, str]] = {}
    fields = {
        "Agency": "name",
        "General": "general_fund",
        "Cash": "cash_fund",
        "Construction": "construction_fund",
        "Federal": "federal_fund",
        "Revolving": "revolving_fund",
        "Total": "all_funds",
    }
    for label, key in fields.items():
        pattern = rf'<span\b[^>]*\bid=["\']{label}Label_(\d+)["\'][^>]*>(.*?)</span>'
        for index_text, fragment in re.findall(pattern, source, re.I | re.S):
            value = unescape(re.sub(r"<[^>]+>", "", fragment)).strip()
            rows.setdefault(int(index_text), {})[key] = value

    agencies = []
    for index in sorted(rows):
        row = rows[index]
        name = row.get("name", "").strip()
        if not name or name.lower().startswith("total"):
            continue
        agency = {"id": str(index), "name": name}
        for key in ("general_fund", "cash_fund", "construction_fund", "federal_fund", "revolving_fund", "all_funds"):
            agency[key] = round(_number(row.get(key, "0")))
        agency["appropriation"] = agency["general_fund"]
        if agency["all_funds"] or agency["general_fund"]:
            agencies.append(agency)
    return agencies


def parse_agency_budget_html(html: str) -> tuple[list[dict], str]:
    parser = TableParser()
    parser.feed(html)
    selected = None
    for table in parser.tables:
        if not table:
            continue
        header = " | ".join(cell.lower() for cell in table[0])
        if all(needle in header for needle in ("agency", "general", "cash", "federal", "total")):
            selected = table
            break
    if selected is None:
        agencies = _parse_agency_budget_spans(html)
        if len(agencies) < 10:
            raise SourceError("Current Fiscal Year Budget page did not contain the expected agency listing")
        fy_match = re.search(r"(?:Fiscal Year|FY)\s*(20\d{2})\s*[-–/]\s*(20\d{2}|\d{2})", html, re.I)
        fy = f"FY{fy_match.group(1)}-{fy_match.group(2)[-2:]}" if fy_match else fiscal_year_label()
        return agencies, fy
    headers = [_normalized_header(value) for value in selected[0]]

    def column(needle: str) -> int | None:
        return next((index for index, header in enumerate(headers) if needle in header), None)

    indices = {name: column(name) for name in ("agency", "general", "cash", "construction", "federal", "revolving", "total")}
    agencies = []
    for index, cells in enumerate(selected[1:], start=1):
        agency_col = indices["agency"]
        if agency_col is None or agency_col >= len(cells):
            continue
        name = cells[agency_col].strip()
        if not name or name.lower().startswith("total"):
            continue

        def amount(key: str) -> int:
            col = indices[key]
            return round(_number(cells[col])) if col is not None and col < len(cells) else 0

        agency = {
            "id": str(index),
            "name": name,
            "general_fund": amount("general"),
            "cash_fund": amount("cash"),
            "construction_fund": amount("construction"),
            "federal_fund": amount("federal"),
            "revolving_fund": amount("revolving"),
            "all_funds": amount("total"),
        }
        agency["appropriation"] = agency["general_fund"]
        if agency["all_funds"] or agency["general_fund"]:
            agencies.append(agency)
    if len(agencies) < 10:
        raise SourceError(f"Current Fiscal Year Budget page yielded only {len(agencies)} agencies")
    fy_match = re.search(r"(?:Fiscal Year|FY)\s*(20\d{2})\s*[-–/]\s*(20\d{2}|\d{2})", html, re.I)
    fy = f"FY{fy_match.group(1)}-{fy_match.group(2)[-2:]}" if fy_match else fiscal_year_label()
    return agencies, fy


# Assembly, fail-safe retention, and optional legacy Sheet upload

def load_previous(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def keep_or_raise(previous: dict, keys: tuple[str, ...], warning: str, warnings: list[str]) -> dict:
    if previous and all(key in previous for key in keys):
        warnings.append(f"{warning} Last known-good values were retained.")
        return {key: previous[key] for key in keys}
    raise SourceError(warning)


def write_json(data: dict, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(output.suffix + ".tmp")
    temporary.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    temporary.replace(output)


def push_to_sheet(data: dict, sheet_id: str, sheet_name: str, credentials_path: str) -> None:
    """Optional compatibility export; the public dashboard no longer needs it."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError("Install google-api-python-client and google-auth for --sheet-id") from exc
    credentials = service_account.Credentials.from_service_account_file(
        credentials_path, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    compact = json.dumps(data, separators=(",", ":"), ensure_ascii=False)
    chunks = [compact[index:index + 40000] for index in range(0, len(compact), 40000)] or ["{}"]
    service.spreadsheets().values().clear(spreadsheetId=sheet_id, range=f"{sheet_name}!A:A").execute()
    service.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range=f"{sheet_name}!A1",
        valueInputOption="RAW",
        body={"values": [[chunk] for chunk in chunks]},
    ).execute()


def build_dashboard(output: Path, target: date | None = None) -> dict:
    previous = load_previous(output)
    warnings: list[str] = []
    sources: dict[str, dict] = {}
    work_dir = Path(tempfile.mkdtemp(prefix="ne-budget-"))

    try:
        oip_path, oip_date, oip_url = fetch_oip(work_dir, target)
        oip = parse_oip(oip_path)
        oip["funds"] = merge_fund_history(oip["funds"], previous.get("funds", []), oip_date)
        sources["cashPool"] = _source("DAS Operating Investment Pool", oip_url, oip_date.strftime("%B %Y"))
    except Exception as exc:
        oip = keep_or_raise(previous, ("macro", "funds"), f"OIP refresh failed: {exc}", warnings)
        sources["cashPool"] = {**previous.get("sources", {}).get("cashPool", {}), "status": "stale"}

    budget_path = None
    status_path = None
    directory_paths: list[Path] = []
    documents = {"status": None, "budget": None, "directories": []}
    try:
        _, legislature_links = official_links(LEGISLATURE_REPORTS_URL)
        legislature_links = [
            Link(CURRENT_GF_STATUS_URL, "Current General Fund Status"),
            *legislature_links,
        ]
        documents = discover_legislature_documents(legislature_links)
    except Exception as exc:
        warnings.append(f"Legislature document discovery failed: {exc}")

    if documents["status"]:
        try:
            status_path = download_document(documents["status"].url, work_dir / "gf-status.pdf", "pdf")
        except Exception as exc:
            warnings.append(f"Current General Fund status download failed: {exc}")
    if documents["budget"]:
        try:
            budget_path = download_document(documents["budget"].url, work_dir / "budget-report.pdf", "pdf")
        except Exception as exc:
            warnings.append(f"Legislative budget report download failed: {exc}")
    for index, link in enumerate(documents["directories"]):
        try:
            directory_paths.append(download_document(link.url, work_dir / f"lfo-directory-{index}.pdf", "pdf"))
        except Exception as exc:
            warnings.append(f"LFO directory volume skipped: {exc}")

    try:
        gf = parse_gf_status(status_path or budget_path, target or date.today())
        gf_link = documents["status"] if status_path else documents["budget"]
        gf_url = gf_link.url if gf_link else CURRENT_GF_STATUS_URL
        gf_period = gf_link.text if gf_link else "Current General Fund Status"
        sources["generalFundStatus"] = _source(
            "Nebraska Legislature General Fund Financial Status",
            gf_url,
            gf_period or "Latest legislative snapshot",
        )
    except Exception as exc:
        retained = keep_or_raise(
            previous,
            ("generalFundStatus", "gfStatusTable", "gfStatusYears"),
            f"General Fund status refresh failed: {exc}",
            warnings,
        )
        gf = {"status": retained["generalFundStatus"], "table": retained["gfStatusTable"], "years": retained["gfStatusYears"]}
        sources["generalFundStatus"] = {**previous.get("sources", {}).get("generalFundStatus", {}), "status": "stale"}

    try:
        _, revenue_links = official_links(REVENUE_REPORTS_URL)
        revenue_date, revenue_url = discover_revenue_release(revenue_links, target)
        revenue_path = download_document(revenue_url, work_dir / "revenue.pdf", "pdf")
        revenue = parse_revenue_text(_pdf_to_text(revenue_path), revenue_date.strftime("%B %Y"))
        if budget_path:
            revenue["nefabForecasts"] = parse_nefab_forecasts(budget_path)
        revenue = merge_monthly_revenue_series(revenue, previous.get("revenue", {}))
        sources["revenue"] = _source(
            "Nebraska Department of Revenue General Fund Receipts", revenue_url, revenue_date.strftime("%B %Y")
        )
    except Exception as exc:
        retained = keep_or_raise(previous, ("revenue",), f"Revenue refresh failed: {exc}", warnings)
        revenue = retained["revenue"]
        sources["revenue"] = {**previous.get("sources", {}).get("revenue", {}), "status": "stale"}

    try:
        agency_html = fetch_html(AGENCY_BUDGET_URL)
        agencies, agency_fy = parse_agency_budget_html(agency_html)
        sources["agencies"] = _source(
            "Nebraska State Spending Current Fiscal Year Budget", AGENCY_BUDGET_URL, agency_fy
        )
    except Exception as exc:
        retained = keep_or_raise(previous, ("agencies",), f"Agency budget refresh failed: {exc}", warnings)
        agencies = retained["agencies"]
        sources["agencies"] = {**previous.get("sources", {}).get("agencies", {}), "status": "stale"}

    descriptions = parse_lfo_directory(directory_paths)
    if len(descriptions) <= 2 and previous.get("fundDescriptions"):
        descriptions = previous["fundDescriptions"]
        warnings.append("LFO fund descriptions could not be refreshed. Last known-good descriptions were retained.")
    if documents.get("directories"):
        sources["fundDirectory"] = _source(
            "Legislative Fiscal Office Fund Directory",
            documents["directories"][0].url,
            "Latest directory edition",
        )

    cr_fund = next((fund for fund in oip.get("funds", []) if fund.get("id") == "11000"), None)
    gf["status"]["cashReserveOipAverageDailyBalance"] = cr_fund.get("balance", 0) if cr_fund else 0
    return {
        "schemaVersion": 2,
        "generatedAt": utc_now(),
        "lastUpdated": {
            "cash": sources.get("cashPool", {}).get("period", "Unknown"),
            "budget": sources.get("generalFundStatus", {}).get("period", "Unknown"),
            "revenue": sources.get("revenue", {}).get("period", "Unknown"),
        },
        "sources": sources,
        "warnings": warnings,
        "population": {
            "value": NEBRASKA_POPULATION,
            "asOf": NEBRASKA_POPULATION_AS_OF,
            "url": CENSUS_POPULATION_URL,
        },
        "macro": oip["macro"],
        "funds": oip["funds"],
        "revenue": revenue,
        "generalFundStatus": gf["status"],
        "gfStatusTable": gf["table"],
        "gfStatusYears": gf["years"],
        "agencies": agencies,
        "fundDescriptions": descriptions,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", default="public/dashboard_data.json")
    parser.add_argument("--month", help="latest report month to include, YYYY-MM")
    parser.add_argument("--sheet-id", help="optional legacy Google Sheet export")
    parser.add_argument("--sheet-name", default="Sheet1")
    parser.add_argument("--credentials-path", default="credentials.json")
    args = parser.parse_args()

    output = Path(args.output)
    dashboard = build_dashboard(output, parse_target_month(args.month))
    previous = load_previous(output)
    if previous:
        old_comparable = {key: value for key, value in previous.items() if key != "generatedAt"}
        new_comparable = {key: value for key, value in dashboard.items() if key != "generatedAt"}
        if old_comparable == new_comparable:
            dashboard["generatedAt"] = previous.get("generatedAt", dashboard["generatedAt"])
    write_json(dashboard, output)
    if args.sheet_id:
        push_to_sheet(dashboard, args.sheet_id, args.sheet_name, args.credentials_path)
    print(
        f"Wrote {output} with {len(dashboard['funds'])} OIP funds, "
        f"{len(dashboard['agencies'])} agencies, and {len(dashboard['warnings'])} warnings."
    )


if __name__ == "__main__":
    main()
